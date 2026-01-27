"""
Data Export Utilities - Excel Generation with README

Creates Excel files with:
- README sheet (comprehensive documentation)
- Data sheets (households, members, errors)
- Professional formatting
- Reference table lookups (code-to-description mappings)
"""
import json
import time
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from src.api.utils.schema_introspection import SchemaIntrospector
from src.core.logging import get_logger
from src.database.engine import engine
from src.database.models import Household, HouseholdMember, QCError
from src.database.reference_models import (
    RefAbawdStatus,
    RefAgencyResponsibility,
    RefCategoricalEligibility,
    RefCitizenshipStatus,
    RefDisability,
    RefDiscovery,
    RefElement,
    RefEmploymentStatusType,
    RefErrorFinding,
    RefExpeditedService,
    RefNature,
    RefRaceEthnicity,
    RefRelationship,
    RefSex,
    RefSnapAffiliation,
    RefStatus,
    RefWorkingIndicator,
    RefWorkRegistration,
)

logger = get_logger(__name__)

# Batch size for streaming large queries (reduces memory usage)
BATCH_SIZE = 5000


class DataExporter:
    """Handles data exports to Excel format with README."""

    # Colors for formatting
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    SECTION_FILL = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    README_TITLE_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
    SECTION_FONT = Font(bold=True, size=12)

    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    def __init__(self, db: Session, data_mapping_path: Path):
        """
        Initialize exporter.

        Args:
            db: Database session
            data_mapping_path: Path to data_mapping.json
        """
        self.db = db
        self.data_mapping = self._load_mapping(data_mapping_path)
        self.introspector = SchemaIntrospector(engine)  # Dynamic schema introspection
        self._reference_cache: dict[str, dict[int, str]] = {}
        self._load_reference_tables()

    def _load_mapping(self, path: Path) -> dict[str, Any]:
        """Load data mapping JSON."""
        with open(path, encoding="utf-8") as f:
            return json.load(f)

    def _load_reference_tables(self):
        """Load all reference tables into memory for efficient lookups."""
        # Define reference table mappings: cache_key -> (model_class, code_field, desc_field, extra_fields)
        ref_tables = [
            ("status", RefStatus, "code", "description", []),
            ("categorical_eligibility", RefCategoricalEligibility, "code", "description", []),
            ("expedited_service", RefExpeditedService, "code", "description", []),
            ("sex", RefSex, "code", "description", []),
            ("snap_affiliation", RefSnapAffiliation, "code", "description", []),
            ("race_ethnicity", RefRaceEthnicity, "code", "description", []),
            ("relationship", RefRelationship, "code", "description", []),
            ("citizenship_status", RefCitizenshipStatus, "code", "description", []),
            ("work_registration", RefWorkRegistration, "code", "description", []),
            ("abawd_status", RefAbawdStatus, "code", "description", []),
            ("disability", RefDisability, "code", "description", []),
            ("working_indicator", RefWorkingIndicator, "code", "description", []),
            ("employment_status", RefEmploymentStatusType, "code", "description", []),
            ("element", RefElement, "code", "description", ["category"]),
            ("nature", RefNature, "code", "description", ["category"]),
            ("agency_responsibility", RefAgencyResponsibility, "code", "description", ["responsibility_type"]),
            ("discovery", RefDiscovery, "code", "description", []),
            ("error_finding", RefErrorFinding, "code", "description", []),
        ]

        for cache_key, model_class, code_field, desc_field, extra_fields in ref_tables:
            try:
                rows = self.db.query(model_class).all()
                # Store as code -> description mapping
                self._reference_cache[cache_key] = {
                    getattr(row, code_field): getattr(row, desc_field)
                    for row in rows
                }
                # Store extra fields if present (e.g., category, responsibility_type)
                for extra in extra_fields:
                    extra_key = f"{cache_key}_{extra}"
                    self._reference_cache[extra_key] = {
                        getattr(row, code_field): getattr(row, extra, None)
                        for row in rows
                    }
            except Exception:
                # Table might not exist or be empty
                self._reference_cache[cache_key] = {}

    def _get_description(self, cache_key: str, code: int | None) -> str | None:
        """Look up description for a code value."""
        if code is None:
            return None
        cache = self._reference_cache.get(cache_key, {})
        return cache.get(code)

    def create_excel_export(
        self,
        tables: list[str] = None,
        fiscal_year: int = None,
        limit: int = None
    ) -> BytesIO:
        """
        Create complete Excel export with README and specified tables.

        Args:
            tables: List of table names to export (default: households, household_members, qc_errors)
            fiscal_year: Optional fiscal year filter
            limit: Optional row limit per table

        Returns:
            BytesIO buffer containing Excel file
        """
        export_start = time.time()

        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Default to core 3 tables if not specified (backward compatibility)
        DEFAULT_TABLES = ["households", "household_members", "qc_errors"]
        tables_to_export = tables or DEFAULT_TABLES

        # 1. Create README sheet (first/default)
        self._create_readme_sheet(wb, fiscal_year, limit, tables_to_export)

        # 2. Create data sheets based on requested tables
        for table_name in tables_to_export:
            # Use specialized methods for core tables (they have reference lookups)
            if table_name == "households":
                self._create_households_sheet(wb, fiscal_year, limit)
            elif table_name == "household_members":
                self._create_members_sheet(wb, fiscal_year, limit)
            elif table_name == "qc_errors":
                self._create_errors_sheet(wb, fiscal_year, limit)
            else:
                # Generic method for custom tables
                self._create_dynamic_sheet(wb, table_name, fiscal_year, limit)

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        total_elapsed = time.time() - export_start
        logger.info(f"Total export time: {total_elapsed:.2f}s")

        return buffer

    def _create_readme_sheet(
        self,
        wb: Workbook,
        fiscal_year: int = None,
        limit: int = None,
        tables: list[str] = None
    ):
        """Create comprehensive README sheet."""
        ws = wb.create_sheet("README", 0)  # First sheet
        ws.sheet_properties.tabColor = "4F81BD"  # Blue tab

        # Default tables if not specified
        if tables is None:
            tables = ["households", "household_members", "qc_errors"]

        row = 1

        # Title
        ws.merge_cells(f"A{row}:E{row}")
        title_cell = ws[f"A{row}"]
        title_cell.value = "SnapAnalyst Data Export - README"
        title_cell.font = self.TITLE_FONT
        title_cell.fill = self.README_TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 25
        row += 2

        # Generation info
        ws[f"A{row}"] = "Generated:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 1

        if fiscal_year:
            ws[f"A{row}"] = "Fiscal Year:"
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = fiscal_year
            row += 1

        if limit:
            ws[f"A{row}"] = "Row Limit:"
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = f"{limit} rows per table"
            row += 1

        row += 1

        # Database Info
        self._add_section(ws, row, "Database Information")
        row += 1

        db_info = self.data_mapping.get("database", {})
        info_items = [
            ("Name", db_info.get("name", "N/A")),
            ("Version", db_info.get("version", "N/A")),
            ("Description", db_info.get("description", "N/A")),
            ("Total Households", db_info.get("total_households", "N/A")),
            ("Fiscal Years", ", ".join(map(str, db_info.get("fiscal_years_available", [])))),
            ("Data Source", db_info.get("data_source", "N/A")),
        ]

        for label, value in info_items:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        row += 1

        # Sheet Descriptions
        self._add_section(ws, row, "Sheets in This Workbook")
        row += 1

        # Build dynamic sheet descriptions
        sheets_info = [("README", "This sheet - Complete documentation and data dictionary")]

        # Add descriptions for each exported table
        table_descriptions = {
            "households": "Household case data (~50k rows per year)",
            "household_members": "Individual household member data (~120k rows)",
            "qc_errors": "Quality control errors/variances (~20k rows)",
        }

        for table_name in tables:
            # Use predefined description or generate generic one
            description = table_descriptions.get(
                table_name,
                f"Custom data table: {table_name}"
            )
            # Convert table name to sheet name format
            sheet_name = table_name.replace("_", " ").title()[:31]
            sheets_info.append((sheet_name, description))

        for sheet_name, description in sheets_info:
            ws[f"A{row}"] = sheet_name
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = description
            row += 1

        row += 1

        # Description Columns Explanation
        self._add_description_columns_info(ws, row)
        row = ws.max_row + 2

        # Table Columns
        for table_name in ["households", "household_members", "qc_errors"]:
            self._add_table_columns(ws, row, table_name)
            row = ws.max_row + 2

        # Code Lookups
        self._add_code_lookups(ws, row)
        row = ws.max_row + 2

        # Relationships
        self._add_relationships(ws, row)

        # Auto-size columns
        for col in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 25
        ws.column_dimensions["C"].width = 50  # Description column wider

    def _add_section(self, ws, row: int, title: str):
        """Add a section header."""
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws[f"A{row}"]
        cell.value = title
        cell.font = self.SECTION_FONT
        cell.fill = self.SECTION_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20

    def _add_description_columns_info(self, ws, start_row: int):
        """Add documentation about the _desc columns added for code lookups."""
        self._add_section(ws, start_row, "Description Columns (Code-to-Text Lookups)")
        row = start_row + 1

        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = "Each coded field has a corresponding '_desc' column with human-readable text from reference tables."
        row += 2

        # Headers
        ws[f"A{row}"] = "Code Column"
        ws[f"B{row}"] = "Description Column"
        ws[f"C{row}"] = "Reference Table"
        ws[f"D{row}"] = "Example Value"
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].font = self.HEADER_FONT
            ws[f"{col}{row}"].fill = self.HEADER_FILL
            ws[f"{col}{row}"].border = self.THIN_BORDER
        row += 1

        # Households columns
        households_mappings = [
            ("status", "status_desc", "ref_status", "Amount correct"),
            ("categorical_eligibility", "categorical_eligibility_desc", "ref_categorical_eligibility", "Unit categorically eligible"),
            ("expedited_service", "expedited_service_desc", "ref_expedited_service", "Not entitled to expedited service"),
        ]

        # Members columns
        members_mappings = [
            ("sex", "sex_desc", "ref_sex", "Female"),
            ("race_ethnicity", "race_ethnicity_desc", "ref_race_ethnicity", "Black or African American (Not Hispanic)"),
            ("relationship_to_head", "relationship_to_head_desc", "ref_relationship", "Head of household"),
            ("citizenship_status", "citizenship_status_desc", "ref_citizenship_status", "US-born citizen"),
            ("snap_affiliation_code", "snap_affiliation_code_desc", "ref_snap_affiliation", "Eligible member entitled to receive benefits"),
            ("disability_indicator", "disability_indicator_desc", "ref_disability", "Not disabled"),
            ("work_registration_status", "work_registration_status_desc", "ref_work_registration", "Work registrant"),
            ("abawd_status", "abawd_status_desc", "ref_abawd_status", "Not an ABAWD"),
            ("working_indicator", "working_indicator_desc", "ref_working_indicator", "Working"),
            ("employment_status_a", "employment_status_a_desc", "ref_employment_status_type", "Employed by other"),
            ("employment_status_b", "employment_status_b_desc", "ref_employment_status_type", "Employed by other"),
        ]

        # QC Errors columns
        errors_mappings = [
            ("element_code", "element_code_desc", "ref_element", "Wages and salaries"),
            ("element_code", "element_category", "ref_element (category)", "earned_income"),
            ("nature_code", "nature_code_desc", "ref_nature", "Unreported source of income"),
            ("nature_code", "nature_category", "ref_nature (category)", "income"),
            ("responsible_agency", "responsible_agency_desc", "ref_agency_responsibility", "Policy incorrectly applied"),
            ("responsible_agency", "agency_responsibility_responsibility_type", "ref_agency_responsibility (type)", "agency"),
            ("discovery_method", "discovery_method_desc", "ref_discovery", "Variance discovered from recipient interview"),
            ("error_finding", "error_finding_desc", "ref_error_finding", "Overissuance"),
        ]

        # Write all mappings grouped by sheet
        ws[f"A{row}"] = "--- Households Sheet ---"
        ws[f"A{row}"].font = Font(bold=True, italic=True)
        row += 1

        for code_col, desc_col, ref_table, example in households_mappings:
            ws[f"A{row}"] = code_col
            ws[f"B{row}"] = desc_col
            ws[f"C{row}"] = ref_table
            ws[f"D{row}"] = example
            for col in ["A", "B", "C", "D"]:
                ws[f"{col}{row}"].border = self.THIN_BORDER
            row += 1

        row += 1
        ws[f"A{row}"] = "--- Members Sheet ---"
        ws[f"A{row}"].font = Font(bold=True, italic=True)
        row += 1

        for code_col, desc_col, ref_table, example in members_mappings:
            ws[f"A{row}"] = code_col
            ws[f"B{row}"] = desc_col
            ws[f"C{row}"] = ref_table
            ws[f"D{row}"] = example
            for col in ["A", "B", "C", "D"]:
                ws[f"{col}{row}"].border = self.THIN_BORDER
            row += 1

        row += 1
        ws[f"A{row}"] = "--- QC_Errors Sheet ---"
        ws[f"A{row}"].font = Font(bold=True, italic=True)
        row += 1

        for code_col, desc_col, ref_table, example in errors_mappings:
            ws[f"A{row}"] = code_col
            ws[f"B{row}"] = desc_col
            ws[f"C{row}"] = ref_table
            ws[f"D{row}"] = example
            for col in ["A", "B", "C", "D"]:
                ws[f"{col}{row}"].border = self.THIN_BORDER
            row += 1

    def _add_table_columns(self, ws, start_row: int, table_name: str):
        """Add table column definitions (from live database introspection)."""
        # Get live table structure from database
        db_table_info = self.introspector.get_table_structure(table_name)

        # Get static metadata for enrichment
        static_table_info = self.data_mapping.get("tables", {}).get(table_name, {})

        # Section header
        display_name = table_name.replace("_", " ").title()
        self._add_section(ws, start_row, f"{display_name} Table - Column Definitions")
        row = start_row + 1

        # Table description (from static metadata)
        ws[f"A{row}"] = "Description:"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = static_table_info.get("description", "N/A")
        row += 1

        # Row Count (from live database)
        ws[f"A{row}"] = "Row Count:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = f"{db_table_info.get('row_count', 0):,}"
        row += 2

        # Column headers
        headers = ["Column", "Type", "Description", "Range", "Example"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.THIN_BORDER

        row += 1

        # Column data (from database + static metadata)
        db_columns = db_table_info.get("columns", {})
        static_columns = static_table_info.get("columns", {})

        for col_name, db_col_info in db_columns.items():
            # Get static metadata for this column if available
            static_col_info = static_columns.get(col_name, {})

            ws[f"A{row}"] = col_name
            ws[f"B{row}"] = db_col_info.get("type", "")  # From database
            ws[f"C{row}"] = static_col_info.get("description", "")  # From static JSON
            ws[f"D{row}"] = str(static_col_info.get("range", ""))  # From static JSON
            ws[f"E{row}"] = str(static_col_info.get("example", ""))  # From static JSON

            # Add borders
            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).border = self.THIN_BORDER

            row += 1

    def _add_code_lookups(self, ws, start_row: int):
        """Add code lookup tables."""
        self._add_section(ws, start_row, "Code Lookup Tables")
        row = start_row + 1

        ws[f"A{row}"] = "Use these tables to decode numeric codes in the data."
        row += 2

        code_lookups = self.data_mapping.get("code_lookups", {})

        for lookup_name, lookup_data in code_lookups.items():
            # Lookup name
            ws[f"A{row}"] = lookup_name.replace("_", " ").title()
            ws[f"A{row}"].font = Font(bold=True, size=11)
            row += 1

            ws[f"A{row}"] = "Description:"
            ws[f"B{row}"] = lookup_data.get("description", "N/A")
            row += 1

            ws[f"A{row}"] = "Source Field:"
            ws[f"B{row}"] = lookup_data.get("source_field", "N/A")
            row += 1

            # Code table headers
            ws[f"A{row}"] = "Code"
            ws[f"B{row}"] = "Description"
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].font = Font(bold=True)
            ws[f"A{row}"].fill = self.SECTION_FILL
            ws[f"B{row}"].fill = self.SECTION_FILL
            row += 1

            # Codes
            for code, description in lookup_data.items():
                if code not in ["description", "source_field"]:
                    ws[f"A{row}"] = str(code)
                    ws[f"B{row}"] = str(description)
                    row += 1

            row += 1

    def _add_relationships(self, ws, start_row: int):
        """Add table relationships."""
        self._add_section(ws, start_row, "Table Relationships")
        row = start_row + 1

        ws[f"A{row}"] = "How tables connect via foreign keys:"
        row += 2

        relationships = self.data_mapping.get("relationships", {})

        for rel_name, rel_info in relationships.items():
            ws[f"A{row}"] = rel_name.replace("_", " ").title()
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

            ws[f"A{row}"] = "Type:"
            ws[f"B{row}"] = rel_info.get("type", "N/A")
            row += 1

            ws[f"A{row}"] = "Description:"
            ws[f"B{row}"] = rel_info.get("description", "N/A")
            row += 1

            ws[f"A{row}"] = "Join:"
            ws[f"B{row}"] = rel_info.get("join", "N/A")
            row += 2

    def _create_households_sheet(self, wb: Workbook, fiscal_year: int = None, limit: int = None):
        """Create households data sheet with reference lookups."""
        start_time = time.time()

        ws = wb.create_sheet("Households")
        ws.sheet_properties.tabColor = "00B050"  # Green

        # Get global filter
        from src.core.filter_manager import get_filter_manager
        filter_manager = get_filter_manager()
        current_filter = filter_manager.get_filter()

        # Query data - ORDER BY case_id to get consistent ordering
        query = self.db.query(Household).order_by(Household.case_id, Household.fiscal_year)

        # Apply global state filter
        if current_filter.state:
            query = query.filter(Household.state_name == current_filter.state)

        # Apply fiscal year filter
        if fiscal_year:
            query = query.filter(Household.fiscal_year == fiscal_year)
        elif current_filter.fiscal_year:
            query = query.filter(Household.fiscal_year == current_filter.fiscal_year)

        # Apply limit
        if limit:
            query = query.limit(limit)

        # Use yield_per for batch processing (streams results instead of loading all into memory)
        query = query.yield_per(BATCH_SIZE)

        # Define columns with optional reference lookups
        # Format: (column_name, ref_cache_key or None)
        column_defs = [
            ("case_id", None),
            ("fiscal_year", None),
            ("state_code", None),
            ("state_name", None),
            ("year_month", None),
            ("status", "status"),  # Add status_desc
            ("snap_benefit", None),
            ("raw_benefit", None),
            ("amount_error", None),
            ("gross_income", None),
            ("net_income", None),
            ("earned_income", None),
            ("unearned_income", None),
            ("certified_household_size", None),
            ("num_elderly", None),
            ("num_children", None),
            ("num_disabled", None),
            ("categorical_eligibility", "categorical_eligibility"),  # Add categorical_eligibility_desc
            ("expedited_service", "expedited_service"),  # Add expedited_service_desc
            ("working_poor_indicator", None),
            ("tanf_indicator", None),
        ]

        # Build headers list (includes _desc columns)
        headers = []
        for col_name, ref_key in column_defs:
            headers.append(col_name)
            if ref_key:
                headers.append(f"{col_name}_desc")

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Data rows - stream results in batches
        row_idx = 2
        row_count = 0
        for household in query:
            col_idx = 1
            for col_name, ref_key in column_defs:
                value = getattr(household, col_name, None)
                ws.cell(row=row_idx, column=col_idx, value=value)
                col_idx += 1

                if ref_key:
                    # Add description column
                    desc = self._get_description(ref_key, value)
                    ws.cell(row=row_idx, column=col_idx, value=desc)
                    col_idx += 1

            row_idx += 1
            row_count += 1

        # Handle empty result set
        if row_count == 0:
            ws["A2"] = "No data available"
        else:
            # Freeze header row
            ws.freeze_panes = "A2"

            # Auto-size columns
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 18

        elapsed = time.time() - start_time
        logger.info(f"Created households sheet with {row_count} rows in {elapsed:.2f}s")

    def _create_members_sheet(self, wb: Workbook, fiscal_year: int = None, limit: int = None):
        """Create household members data sheet with reference lookups."""
        start_time = time.time()

        ws = wb.create_sheet("Members")
        ws.sheet_properties.tabColor = "FFC000"  # Orange

        # Get global filter
        from src.core.filter_manager import get_filter_manager
        filter_manager = get_filter_manager()
        current_filter = filter_manager.get_filter()

        # Query data - SELECT state_name directly from joined Household to avoid lazy loading
        query = self.db.query(
            HouseholdMember,
            Household.state_name
        ).join(
            Household,
            (HouseholdMember.case_id == Household.case_id) &
            (HouseholdMember.fiscal_year == Household.fiscal_year)
        ).order_by(HouseholdMember.case_id, HouseholdMember.fiscal_year, HouseholdMember.member_number)

        # Apply global state filter
        if current_filter.state:
            query = query.filter(Household.state_name == current_filter.state)

        # Apply fiscal year filter
        if fiscal_year:
            query = query.filter(Household.fiscal_year == fiscal_year)
        elif current_filter.fiscal_year:
            query = query.filter(Household.fiscal_year == current_filter.fiscal_year)

        # Apply limit
        if limit:
            query = query.limit(limit)

        # Use yield_per for batch processing
        query = query.yield_per(BATCH_SIZE)

        # Define columns with optional reference lookups
        # Format: (column_name, ref_cache_key or None, is_from_join)
        column_defs = [
            ("case_id", None, False),
            ("fiscal_year", None, False),
            ("state_name", None, True),  # From joined Household
            ("member_number", None, False),
            ("age", None, False),
            ("sex", "sex", False),  # Add sex_desc
            ("race_ethnicity", "race_ethnicity", False),  # Add race_ethnicity_desc
            ("relationship_to_head", "relationship", False),  # Add relationship_desc
            ("citizenship_status", "citizenship_status", False),  # Add citizenship_status_desc
            ("snap_affiliation_code", "snap_affiliation", False),  # Add snap_affiliation_desc
            ("disability_indicator", "disability", False),  # Add disability_desc
            ("work_registration_status", "work_registration", False),  # Add work_registration_desc
            ("abawd_status", "abawd_status", False),  # Add abawd_status_desc
            ("working_indicator", "working_indicator", False),  # Add working_indicator_desc
            ("employment_status_a", "employment_status", False),  # Add employment_status_a_desc
            ("employment_status_b", "employment_status", False),  # Add employment_status_b_desc
            ("wages", None, False),
            ("self_employment_income", None, False),
            ("social_security", None, False),
            ("ssi", None, False),
            ("unemployment", None, False),
            ("tanf", None, False),
            ("child_support", None, False),
            ("total_income", None, False),
        ]

        # Build headers list (includes _desc columns)
        headers = []
        for col_name, ref_key, _ in column_defs:
            headers.append(col_name)
            if ref_key:
                headers.append(f"{col_name}_desc")

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Data rows - stream results in batches (each result is a tuple: member, state_name)
        row_idx = 2
        row_count = 0
        for member, state_name in query:
            col_idx = 1
            for col_name, ref_key, is_from_join in column_defs:
                # Get value
                value = state_name if is_from_join and col_name == "state_name" else getattr(member, col_name, None)

                ws.cell(row=row_idx, column=col_idx, value=value)
                col_idx += 1

                if ref_key:
                    # Add description column
                    desc = self._get_description(ref_key, value)
                    ws.cell(row=row_idx, column=col_idx, value=desc)
                    col_idx += 1

            row_idx += 1
            row_count += 1

        # Handle empty result set
        if row_count == 0:
            ws["A2"] = "No data available"
        else:
            # Freeze header row
            ws.freeze_panes = "A2"

            # Auto-size columns
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 18

        elapsed = time.time() - start_time
        logger.info(f"Created members sheet with {row_count} rows in {elapsed:.2f}s")

    def _create_errors_sheet(self, wb: Workbook, fiscal_year: int = None, limit: int = None):
        """Create QC errors data sheet with reference lookups."""
        start_time = time.time()

        ws = wb.create_sheet("QC_Errors")
        ws.sheet_properties.tabColor = "FF0000"  # Red

        # Get global filter
        from src.core.filter_manager import get_filter_manager
        filter_manager = get_filter_manager()
        current_filter = filter_manager.get_filter()

        # Query data - SELECT state_name directly from joined Household to avoid lazy loading
        query = self.db.query(
            QCError,
            Household.state_name
        ).join(
            Household,
            (QCError.case_id == Household.case_id) &
            (QCError.fiscal_year == Household.fiscal_year)
        ).order_by(QCError.case_id, QCError.fiscal_year, QCError.error_number)

        # Apply global state filter
        if current_filter.state:
            query = query.filter(Household.state_name == current_filter.state)

        # Apply fiscal year filter
        if fiscal_year:
            query = query.filter(Household.fiscal_year == fiscal_year)
        elif current_filter.fiscal_year:
            query = query.filter(Household.fiscal_year == current_filter.fiscal_year)

        # Apply limit
        if limit:
            query = query.limit(limit)

        # Use yield_per for batch processing
        query = query.yield_per(BATCH_SIZE)

        # Define columns with optional reference lookups
        # Format: (column_name, ref_cache_key or None, is_from_join, extra_ref_keys)
        # extra_ref_keys allows adding additional columns like category
        column_defs = [
            ("case_id", None, False, []),
            ("fiscal_year", None, False, []),
            ("state_name", None, True, []),  # From joined Household
            ("error_number", None, False, []),
            ("element_code", "element", False, ["element_category"]),  # Add element_desc and element_category
            ("nature_code", "nature", False, ["nature_category"]),  # Add nature_desc and nature_category
            ("responsible_agency", "agency_responsibility", False, ["agency_responsibility_responsibility_type"]),  # Add agency_desc and responsibility_type
            ("error_amount", None, False, []),
            ("discovery_method", "discovery", False, []),  # Add discovery_desc
            ("verification_status", None, False, []),
            ("occurrence_date", None, False, []),
            ("time_period", None, False, []),
            ("error_finding", "error_finding", False, []),  # Add error_finding_desc
        ]

        # Build headers list (includes _desc and extra columns)
        headers = []
        for col_name, ref_key, _, extra_keys in column_defs:
            headers.append(col_name)
            if ref_key:
                headers.append(f"{col_name}_desc")
                for extra_key in extra_keys:
                    # Convert cache key to header name (e.g., 'element_category' -> 'element_category')
                    headers.append(extra_key.replace("_", "_", 1))

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Data rows - stream results in batches (each result is a tuple: error, state_name)
        row_idx = 2
        row_count = 0
        for error, state_name in query:
            col_idx = 1
            for col_name, ref_key, is_from_join, extra_keys in column_defs:
                # Get value
                value = state_name if is_from_join and col_name == "state_name" else getattr(error, col_name, None)

                ws.cell(row=row_idx, column=col_idx, value=value)
                col_idx += 1

                if ref_key:
                    # Add description column
                    desc = self._get_description(ref_key, value)
                    ws.cell(row=row_idx, column=col_idx, value=desc)
                    col_idx += 1

                    # Add extra columns (category, responsibility_type, etc.)
                    for extra_key in extra_keys:
                        extra_val = self._get_description(extra_key, value)
                        ws.cell(row=row_idx, column=col_idx, value=extra_val)
                        col_idx += 1

            row_idx += 1
            row_count += 1

        # Handle empty result set
        if row_count == 0:
            ws["A2"] = "No data available"
        else:
            # Freeze header row
            ws.freeze_panes = "A2"

            # Auto-size columns
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 20

        elapsed = time.time() - start_time
        logger.info(f"Created errors sheet with {row_count} rows in {elapsed:.2f}s")

    def _create_dynamic_sheet(
        self,
        wb: Workbook,
        table_name: str,
        fiscal_year: int = None,
        limit: int = None
    ):
        """
        Create a sheet for any table dynamically using schema introspection.

        Works for both core tables and custom snap_* tables.

        Args:
            wb: Workbook to add sheet to
            table_name: Name of the table to export
            fiscal_year: Optional fiscal year filter
            limit: Optional row limit
        """
        from sqlalchemy import text

        from src.core.filter_manager import get_filter_manager

        # Get table structure from database
        try:
            table_info = self.introspector.get_table_structure(table_name)
        except Exception as e:
            logger.warning(f"Could not get structure for table {table_name}: {e}")
            return

        # Create sheet with nice name (Excel has 31 char limit)
        sheet_name = table_name.replace("_", " ").title()[:31]
        ws = wb.create_sheet(sheet_name)

        # Get columns from database
        columns = list(table_info.get("columns", {}).keys())
        if not columns:
            ws["A1"] = f"No columns found for {table_name}"
            return

        # Build query with filters
        query = f"SELECT {', '.join(columns)} FROM {table_name}"
        where_clauses = []
        params = {}

        # Add fiscal_year filter if column exists
        if "fiscal_year" in columns and fiscal_year:
            where_clauses.append("fiscal_year = :fiscal_year")
            params["fiscal_year"] = fiscal_year

        # Add state filter if column exists
        filter_manager = get_filter_manager()
        current_filter = filter_manager.get_filter()
        if "state_name" in columns and current_filter.state:
            where_clauses.append("state_name = :state_name")
            params["state_name"] = current_filter.state

        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)

        # Add ordering for consistency (if case_id exists)
        if "case_id" in columns:
            order_cols = ["case_id"]
            if "fiscal_year" in columns:
                order_cols.append("fiscal_year")
            query += f" ORDER BY {', '.join(order_cols)}"

        if limit:
            query += f" LIMIT {limit}"

        # Execute query
        try:
            result = self.db.execute(text(query), params)
            rows = result.fetchall()
        except Exception as e:
            logger.error(f"Could not query table {table_name}: {e}")
            ws["A1"] = f"Error querying {table_name}: {str(e)}"
            return

        if not rows:
            ws["A1"] = "No data available"
            return

        # Write headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-size columns
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        logger.info(f"Created sheet for table {table_name} with {len(rows)} rows")
